import os
from pathlib import Path
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from inventario.models import Equipo


def limpiar(v):
    if v is None:
        return ''
    return str(v).strip()


class Command(BaseCommand):
    help = 'Importa equipos desde los archivos Excel'

    def add_arguments(self, parser):
        parser.add_argument('--force', action='store_true', help='Re-importar incluso si ya hay datos')

    def handle(self, *args, **options):
        data_dir = Path(__file__).resolve().parent.parent.parent.parent / 'data'

        if Equipo.objects.exists() and not options['force']:
            self.stdout.write(f'Ya hay {Equipo.objects.count()} equipos. Usa --force para re-importar.')
            return

        Equipo.objects.all().delete()

        errores = []
        def _seguro(nombre, fn):
            try:
                fn()
            except Exception as e:
                errores.append(f'{nombre}: {e}')
                self.stdout.write(self.style.ERROR(f'  Error en {nombre}: {e}'))

        _seguro('RX', lambda: self._importar_rx(str(data_dir / 'Estado de los RX Provincia x marca .xlsx')))
        _seguro('USD Marcas', lambda: self._importar_usd_marcas(str(data_dir / 'Estado de los USD x Marcas.xlsx')))
        _seguro('USD Municipios', lambda: self._importar_usd_municipios(str(data_dir / 'Estado de los USD x Municipios.xlsx')))
        _seguro('Plan Mtto', lambda: self._importar_plan_mtto(str(data_dir / 'Plan de Mtto.xlsx')))
        _seguro('Conteo', lambda: self.stdout.write(f'  Parcial: {Equipo.objects.count()} equipos hasta ahora'))

        if errores:
            self.stdout.write(self.style.ERROR(f'{len(errores)} archivo(s) con errores: {"; ".join(errores)}'))

        total = Equipo.objects.count()
        self.stdout.write(self.style.SUCCESS(f'{total} equipos importados.'))

    def _importar_rx(self, path):
        wb = load_workbook(path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                Equipo.objects.create(
                    municipio=limpiar(row[0]),
                    tipo='RX',
                    unidad_salud=limpiar(row[2]),
                    denominacion=limpiar(row[3]),
                    marca=limpiar(row[4]),
                    modelo=limpiar(row[5]),
                    numero_serie=limpiar(row[6]),
                    observaciones=limpiar(row[7]),
                    estado=limpiar(row[8]),
                    fuente='Estado de los RX Provincia x marca',
                )
        self.stdout.write(f'  RX: {wb.sheetnames}')

    def _importar_usd_marcas(self, path):
        wb = load_workbook(path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                Equipo.objects.create(
                    municipio=limpiar(row[0]),
                    tipo='USD',
                    unidad_salud=limpiar(row[1]),
                    marca=limpiar(row[2]),
                    modelo=limpiar(row[3]),
                    numero_serie=limpiar(row[4]),
                    observaciones=limpiar(row[5]),
                    estado=limpiar(row[6]),
                    fuente='Estado de los USD x Marcas',
                )
        self.stdout.write(f'  USD Marcas: {wb.sheetnames}')

    def _mapear_columnas(self, headers):
        """Detecta el índice de cada columna según el encabezado."""
        cols = {}
        for i, h in enumerate(headers):
            if h is None:
                continue
            h = str(h).strip().lower()
            if 'municipio' in h:
                cols['municipio'] = i
            elif 'unidad de salud' in h:
                cols['unidad_salud'] = i
            elif h == 'local':
                cols['local'] = i
            elif 'servicio' in h and 'especialidad' not in h:
                cols['servicio'] = i
            elif 'especialidad' in h:
                cols['especialidad'] = i
            elif h == 'marca':
                cols['marca'] = i
            elif h == 'modelo':
                cols['modelo'] = i
            elif 'serie' in h:
                cols['numero_serie'] = i
            elif 'obs' in h:
                cols['observaciones'] = i
            elif 'estado' in h:
                cols['estado'] = i
        return cols

    def _importar_usd_municipios(self, path):
        wb = load_workbook(path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = list(rows[0])
            cols = self._mapear_columnas(headers)
            for row in rows[1:]:
                if not row[0]:
                    continue
                equipo_data = {
                    'tipo': 'USD',
                    'fuente': 'Estado de los USD x Municipios',
                }
                for campo, idx in cols.items():
                    if idx < len(row) and row[idx] is not None:
                        if campo == 'especialidad':
                            continue
                        equipo_data[campo] = limpiar(row[idx])
                Equipo.objects.create(**equipo_data)
        self.stdout.write(f'  USD Municipios: {wb.sheetnames}')

    def _importar_plan_mtto(self, path):
        wb = load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            Equipo.objects.create(
                tipo='OTRO',
                unidad_salud=limpiar(row[0]),
                denominacion=limpiar(row[1]),
                marca=limpiar(row[2]),
                modelo=limpiar(row[3]),
                numero_serie=limpiar(row[4]),
                estado=limpiar(row[5]),
                frecuencia=limpiar(row[6]),
                fuente='Plan de Mtto',
            )
        self.stdout.write('  Plan de Mtto')
